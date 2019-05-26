from openpyxl import load_workbook
from openpyxl import Workbook
import math

# 28列
###########################################
def makeARowWithBefore(given_row, year):
    result_list = []
    result_list.append(given_row[0].value)
    result_list.append(given_row[1].value)
    result_list.append(given_row[2].value)
    result_list.append(year)

    for i in range(24):
        result_list.append(0)

    return result_list

def makeARowFromId(id, year):
    print('in makeARowFromId, id=', id)
    result = []
    result.append(id[0])
    result.append(id[1])
    result.append(id[2])
    result.append(year)
    for i in range(24):
        result.append(0)

    return result
###########################################
def getFromExcel(given_row):
    temp = []
    i = 0
    for i in range(len(given_row)):
        temp.append(row[i].value)

    return temp

src_dir = './files/原始数据.xlsx'
wb = load_workbook(filename=src_dir)

ws = wb['Sheet1']
isFirstRow = True
Year = 2000
expectedYear = 2000

dest_dir = './files/result.xlsx'
dest_wb = Workbook()
work_sheet = dest_wb.active
var = 0

count = 0
lastId = ''
for row in ws:
    if row[3].value == None:
        continue

    expectedYear = Year + (var % 16)
    print('----->row,', row[3].value)
    print('expectedYear=', expectedYear)
    # 处理第一行
    if isFirstRow:
        isFirstRow = False
        work_sheet.append(getFromExcel(row))
        continue

    if row[3].value == expectedYear:
        work_sheet.append(getFromExcel(row))

        var = var + 1
        expectedYear = Year + (var % 16)
        lastId = str(row[0].value) + '.' + str(row[1].value) + '.' + str(row[2].value)
    else:
        currentId = str(row[0].value) + '.' + str(row[1].value) + '.' + str(row[2].value)
        # 不跨ID
        if lastId == currentId:
            print('不跨IDexpectedyear=', expectedYear)
            print('不跨IDrow3.value=', row[3].value)
            for y in range(expectedYear, row[3].value):
                newRow = makeARowWithBefore(row, y)
                work_sheet.append(newRow)
                var = var + 1

            work_sheet.append(getFromExcel(row))
            var = var + 1
            expectedYear = Year + (var % 16)
        else:
            print('跨ID expectedYear=', expectedYear)
            if expectedYear != 2000:
                for y in range(expectedYear, 2016):
                    var = var + 1
                    work_sheet.append(makeARowFromId(lastId.split('.'), y))
            print("row[3].value=", row[3].value)
            for y in range(2000, row[3].value):
                work_sheet.append(makeARowWithBefore(row, y))
                var = var + 1
            work_sheet.append(getFromExcel(row))
            var = var + 1
            lastId = currentId

    count = count + 1


dest_wb.save(dest_dir)
